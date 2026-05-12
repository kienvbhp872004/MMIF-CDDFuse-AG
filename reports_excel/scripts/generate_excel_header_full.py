import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Create new workbook
wb = Workbook()
ws = wb.active
ws.title = "Báo cáo Tiến độ"

# Add Header Info
header_info = [
    ("ĐẠI HỌC BÁCH KHOA HÀ NỘI", "", "", "", "", "", ""),
    ("TRƯỜNG CÔNG NGHỆ THÔNG TIN VÀ TRUYỀN THÔNG", "", "", "", "", "", ""),
    ("", "", "", "", "", "", ""),
    ("BÁO CÁO TIẾN ĐỘ ĐỒ ÁN TỐT NGHIỆP HỆ CỬ NHÂN", "", "", "", "", "", ""),
    ("KỲ 20252", "", "", "", "", "", ""),
    ("", "", "", "", "", "", ""),
    ("Thông tin về sinh viên", "", "", "", "", "", ""),
    ("Họ và tên sinh viên:", "", "Đỗ Trung Kiên", "", "MSSV:", "20224869", ""),
    ("Điện thoại liên lạc:", "", "0913619590", "", "Lớp:", "Khoa học Máy tính 06", ""),
    ("Email:", "", "kien.dt224869@sis.hust.edu.vn", "", "Mã lớp:", "762000", ""),
    ("Thông tin giáo viên hướng dẫn", "", "", "", "", "", ""),
    ("Họ và tên GVHD:", "", "Phạm Đăng Hải, Phạm Văn Hải", "", "", "", ""),
    ("Đồ án được thực hiện tại:", "", "Trường Công Nghệ thông tin và Truyền thông", "", "", "", ""),
    ("Thời gian làm ĐATN:", "", "Từ ngày 20/02/2026 đến ngày 24/06/2026", "", "", "", ""),
    ("Tên đề tài:", "", "Cải tiến các phương pháp tổng hợp ảnh y tế đa phương thức nhằm nâng cao chất lượng chẩn đoán y khoa.", "", "", "", ""),
    ("", "", "", "", "", "", ""),
    ("TIẾN ĐỘ THỰC HIỆN ĐẾN THỜI ĐIỂM HIỆN TẠI (24/04/2026)", "", "", "", "", "", ""),
    ("STT", "Tuần / Thời gian", "Nội dung công việc", "Chi tiết công việc theo Đề cương", "Trạng thái", "Tiến độ (%)", "Kết quả thực tế đạt được")
]

for row in header_info:
    ws.append(row)

# Styling header section
bold_font = Font(bold=True)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, italic=True, size=12)

ws['A1'].font = bold_font
ws['A2'].font = bold_font
ws['A4'].font = title_font
ws['A4'].alignment = Alignment(horizontal="center")
ws.merge_cells('A4:G4')
ws['A5'].font = bold_font
ws['A5'].alignment = Alignment(horizontal="center")
ws.merge_cells('A5:G5')

ws['A7'].font = section_font
ws['A11'].font = section_font
ws['A17'].font = title_font
ws['A17'].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells('A17:G17')

ws.merge_cells('A8:B8')
ws.merge_cells('C8:D8')
ws.merge_cells('A9:B9')
ws.merge_cells('C9:D9')
ws.merge_cells('A10:B10')
ws.merge_cells('C10:D10')

ws.merge_cells('A12:B12')
ws.merge_cells('C12:E12')
ws.merge_cells('A13:B13')
ws.merge_cells('C13:G13')
ws.merge_cells('A14:B14')
ws.merge_cells('C14:G14')
ws.merge_cells('A15:B15')
ws.merge_cells('C15:G15')

# Make the labels bold
for row_idx in range(8, 11):
    ws[f'A{row_idx}'].font = bold_font
    ws[f'E{row_idx}'].font = bold_font
    
for row_idx in range(12, 16):
    ws[f'A{row_idx}'].font = bold_font

# Set wrap text for project title
ws['C15'].alignment = Alignment(wrap_text=True)
ws.row_dimensions[15].height = 30

# Append data
data = [
    [1, "Tuần 1-4\n(20/02 - 19/03)", "Nội dung 1: Khảo sát tài liệu và triển khai thử nghiệm các mô hình SOTA", "- Nghiên cứu Image Fusion trong y khoa, lập bảng tổng hợp các phương thức ảnh (MRI, PET, SPECT, CT).\n- Cài đặt và kiểm thử bộ Metrics Suite tự động tính 7 chỉ số (FMI, SCD, Qabf, Qy, SSIM, Entropy, Piella).\n- Cài đặt chạy thử các thuật toán phân rã (Multi-scale Transform, Feature Decomposition, Diffusion-based).", "Hoàn thành", "100%", "Đã khảo sát xong các bài báo khoa học. Hoàn thiện bộ Metrics Suite chạy ổn định bằng Python. Đã chạy thử nghiệm thành công các mô hình SOTA."],
    [2, "Tuần 5-6\n(20/03 - 02/04)", "Nội dung 2: Xây dựng hệ thống đánh giá toàn diện và đề xuất cải tiến", "- Cài đặt và chạy thuật toán SOTA: DWT, NSST, Laplacian Pyramid, DAF, PSFusion.\n- Phân tích nhược điểm của từng SOTA (đặc biệt về khả năng bảo toàn đường biên).\n- Lập bảng gap analysis làm cơ sở đề xuất hướng cải tiến.", "Hoàn thành", "100%", "Đã thiết lập xong môi trường chạy, giải quyết các xung đột thư viện (NumPy, PyTorch). Đánh giá nhược điểm và xác định cần tập trung bảo toàn biên (Edge-preserving)."],
    [3, "Tuần 7-8\n(03/04 - 16/04)", "Nội dung 3.1: Xác định hướng cải tiến & Thiết kế Hybrid Fusion", "- Từ kết quả gap analysis, xác định hướng cải tiến cụ thể về edge-preserving filter.\n- Thiết kế kiến trúc block diagram cho Hybrid Fusion Model, chỉ rõ module cải tiến.", "Hoàn thành", "100%", "Đã phác thảo xong kiến trúc và phương pháp tiếp cận mới để tăng cường chất lượng tổng hợp ảnh, giữ được chi tiết biên tốt hơn."],
    [4, "Tuần 9-10\n(17/04 - 30/04)", "Nội dung 3.2: Triển khai luật tổng hợp cải tiến", "- Cài đặt module Guided Filter Enhancement cải tiến.\n- Tích hợp vào pipeline fusion và chạy thử trên 10 cặp ảnh y tế đa kênh.\n- Đo lường và tinh chỉnh tham số để tối ưu các chỉ số (Qabf, Qy, FMI).", "Đang thực hiện", "80%", "Đang code các mô hình và thuật toán lấy mẫu (ví dụ: tối ưu DDIM cho các mô hình Diffusion). Kết quả đánh giá sơ bộ rất khả quan."],
    [5, "Tuần 11-15\n(01/05 - 04/06)", "Nội dung 4: Thử nghiệm thực tế và Đánh giá thống kê diện rộng", "- Tải và chuẩn hóa dataset (Harvard Atlas, BRATS, TCIA).\n- Chạy đánh giá toàn diện các phương pháp đề xuất và SOTA trên các tập dữ liệu lớn.\n- Kiểm định thống kê và thực hiện Ablation Study.", "Chưa tới hạn", "30%", "Đã bắt đầu xử lý Dataset Harvard Medical Image Fusion (PET, SPECT, CT). Sẵn sàng cho chạy thực nghiệm diện rộng."],
    [6, "Tuần 16-17\n(05/06 - 18/06)", "Nội dung 5: Viết báo cáo học thuật và Đóng gói sản phẩm", "- Setup cấu trúc báo cáo bằng LaTeX và hoàn thành 5 chương.\n- Đóng gói code, viết README, hoàn thiện báo cáo và chuẩn bị slide bảo vệ.", "Đang thực hiện", "25%", "Đã setup LaTeX (report_image_fusion.tex), hoàn thiện phần công thức toán học và phân tích cấu trúc mô hình. Đang cập nhật kết quả song song."]
]

for row in data:
    ws.append(row)

# Style the table
table_start_row = 18
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
header_font_white = Font(color="FFFFFF", bold=True, size=11)
border = Border(
    left=Side(style='thin', color="000000"),
    right=Side(style='thin', color="000000"),
    top=Side(style='thin', color="000000"),
    bottom=Side(style='thin', color="000000")
)

# Table headers
for col in range(1, 8):
    cell = ws.cell(row=table_start_row, column=col)
    cell.fill = header_fill
    cell.font = header_font_white
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Table data
for row in range(table_start_row + 1, table_start_row + 1 + len(data)):
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        # Center align STT, Status, Percent
        if col in [1, 5, 6]:
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

# Set column widths
ws.column_dimensions["A"].width = 6    # STT
ws.column_dimensions["B"].width = 18   # Tuan
ws.column_dimensions["C"].width = 30   # Noi dung cv
ws.column_dimensions["D"].width = 45   # Chi tiet
ws.column_dimensions["E"].width = 15   # Trang thai
ws.column_dimensions["F"].width = 12   # Tien do
ws.column_dimensions["G"].width = 40   # Ket qua

# Save
wb.save("Bao_cao_tien_do_HoanChinh_Do_Trung_Kien.xlsx")
