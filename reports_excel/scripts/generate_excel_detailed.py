import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Create detailed progress data mapped to the project form timeline
# Start date: 20/02/2026. Current date: ~24/04/2026 (Week 9-10)
data = [
    {
        "STT": 1,
        "Tuần / Thời gian": "Tuần 1-4\n(20/02 - 19/03)",
        "Nội dung công việc": "Nội dung 1: Khảo sát tài liệu và triển khai thử nghiệm các mô hình SOTA",
        "Chi tiết công việc theo Đề cương": "- Nghiên cứu Image Fusion trong y khoa, lập bảng tổng hợp các phương thức ảnh (MRI, PET, SPECT, CT).\n- Cài đặt và kiểm thử bộ Metrics Suite tự động tính 7 chỉ số (FMI, SCD, Qabf, Qy, SSIM, Entropy, Piella).\n- Cài đặt chạy thử các thuật toán phân rã (Multi-scale Transform, Feature Decomposition, Diffusion-based).",
        "Trạng thái": "Hoàn thành",
        "Tiến độ (%)": "100%",
        "Kết quả thực tế đạt được": "Đã khảo sát xong các bài báo khoa học. Hoàn thiện bộ Metrics Suite chạy ổn định bằng Python. Đã chạy thử nghiệm thành công các mô hình SOTA."
    },
    {
        "STT": 2,
        "Tuần / Thời gian": "Tuần 5-6\n(20/03 - 02/04)",
        "Nội dung công việc": "Nội dung 2: Xây dựng hệ thống đánh giá toàn diện và đề xuất cải tiến",
        "Chi tiết công việc theo Đề cương": "- Cài đặt và chạy thuật toán SOTA: DWT, NSST, Laplacian Pyramid, DAF, PSFusion.\n- Phân tích nhược điểm của từng SOTA (đặc biệt về khả năng bảo toàn đường biên).\n- Lập bảng gap analysis làm cơ sở đề xuất hướng cải tiến.",
        "Trạng thái": "Hoàn thành",
        "Tiến độ (%)": "100%",
        "Kết quả thực tế đạt được": "Đã thiết lập xong môi trường chạy, giải quyết các xung đột thư viện (NumPy, PyTorch). Đánh giá nhược điểm và xác định cần tập trung bảo toàn biên (Edge-preserving)."
    },
    {
        "STT": 3,
        "Tuần / Thời gian": "Tuần 7-8\n(03/04 - 16/04)",
        "Nội dung công việc": "Nội dung 3.1: Xác định hướng cải tiến & Thiết kế Hybrid Fusion",
        "Chi tiết công việc theo Đề cương": "- Từ kết quả gap analysis, xác định hướng cải tiến cụ thể về edge-preserving filter.\n- Thiết kế kiến trúc block diagram cho Hybrid Fusion Model, chỉ rõ module cải tiến.",
        "Trạng thái": "Hoàn thành",
        "Tiến độ (%)": "100%",
        "Kết quả thực tế đạt được": "Đã phác thảo xong kiến trúc và phương pháp tiếp cận mới để tăng cường chất lượng tổng hợp ảnh, giữ được chi tiết biên tốt hơn."
    },
    {
        "STT": 4,
        "Tuần / Thời gian": "Tuần 9-10\n(17/04 - 30/04)",
        "Nội dung công việc": "Nội dung 3.2: Triển khai luật tổng hợp cải tiến",
        "Chi tiết công việc theo Đề cương": "- Cài đặt module Guided Filter Enhancement cải tiến.\n- Tích hợp vào pipeline fusion và chạy thử trên 10 cặp ảnh y tế đa kênh.\n- Đo lường và tinh chỉnh tham số để tối ưu các chỉ số (Qabf, Qy, FMI).",
        "Trạng thái": "Đang thực hiện",
        "Tiến độ (%)": "80%",
        "Kết quả thực tế đạt được": "Đang code các mô hình và thuật toán lấy mẫu (ví dụ: tối ưu DDIM cho các mô hình Diffusion). Kết quả đánh giá sơ bộ rất khả quan."
    },
    {
        "STT": 5,
        "Tuần / Thời gian": "Tuần 11-15\n(01/05 - 04/06)",
        "Nội dung công việc": "Nội dung 4: Thử nghiệm thực tế và Đánh giá thống kê diện rộng",
        "Chi tiết công việc theo Đề cương": "- Tải và chuẩn hóa dataset (Harvard Atlas, BRATS, TCIA).\n- Chạy đánh giá toàn diện các phương pháp đề xuất và SOTA trên các tập dữ liệu lớn.\n- Kiểm định thống kê và thực hiện Ablation Study.",
        "Trạng thái": "Đang thực hiện",
        "Tiến độ (%)": "30%",
        "Kết quả thực tế đạt được": "Đã bắt đầu xử lý Dataset Harvard Medical Image Fusion (PET, SPECT, CT). Sẵn sàng cho chạy thực nghiệm diện rộng."
    },
    {
        "STT": 6,
        "Tuần / Thời gian": "Tuần 16-17\n(05/06 - 18/06)",
        "Nội dung công việc": "Nội dung 5: Viết báo cáo học thuật và Đóng gói sản phẩm",
        "Chi tiết công việc theo Đề cương": "- Setup cấu trúc báo cáo bằng LaTeX và hoàn thành 5 chương.\n- Đóng gói code, viết README, hoàn thiện báo cáo và chuẩn bị slide bảo vệ.",
        "Trạng thái": "Đang thực hiện",
        "Tiến độ (%)": "25%",
        "Kết quả thực tế đạt được": "Đã setup LaTeX (report_image_fusion.tex), hoàn thiện phần công thức toán học và phân tích cấu trúc mô hình. Đang cập nhật kết quả song song."
    }
]

df = pd.DataFrame(data)

file_name = "Bao_cao_tien_do_ChiTiet_Do_Trung_Kien.xlsx"
with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Báo cáo Tiến độ", index=False)
    
    workbook = writer.book
    worksheet = writer.sheets["Báo cáo Tiến độ"]
    
    # Styling details
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    for cell in worksheet["1:1"]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            # Center align STT, Status, Percent
            if cell.column_letter in ['A', 'E', 'F']:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            
    # Set column widths appropriately
    worksheet.column_dimensions["A"].width = 5   # STT
    worksheet.column_dimensions["B"].width = 18  # Tuan
    worksheet.column_dimensions["C"].width = 30  # Noi dung cv
    worksheet.column_dimensions["D"].width = 50  # Chi tiet
    worksheet.column_dimensions["E"].width = 15  # Trang thai
    worksheet.column_dimensions["F"].width = 12  # Tien do
    worksheet.column_dimensions["G"].width = 40  # Ket qua thuc te
    
print(f"Thành công: Đã tạo file {file_name}")
